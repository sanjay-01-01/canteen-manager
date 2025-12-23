from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum , F
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from datetime import date, timedelta
import datetime # Date parsing ke liye zaroori
from dateutil.relativedelta import relativedelta
import csv
import json
from datetime import date
from .models import Canteen, Expense, Staff, SalaryPayment, StaffLeave, DailyEntry
from .forms import SalaryPaymentForm, ExpenseForm, DailyEntryForm , ConsumptionForm
import openpyxl 
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter  # <--- Ye line add karein
from .forms import StaffLeaveForm
import calendar
from django.db.models import Count, Q
from .forms import StaffForm # Upar import check karein
from django.utils import timezone




# ==========================================
# 1. डैशबोर्ड (Home Dashboard)
# ==========================================
# management/views.py

@login_required
def home_dashboard(request):
    # --- 1. Date Filter Logic (Professional) ---
    filter_date_str = request.GET.get('filter_date') # HTML se date layega (YYYY-MM)
    
    if filter_date_str:
        # Agar user ne filter lagaya hai
        try:
            year, month = map(int, filter_date_str.split('-'))
            today = date(year, month, 1) # Us mahine ki 1st tarikh
        except ValueError:
            today = date.today()
    else:
        # Default: Aaj ka din
        today = date.today()

    # Is mahine ki range nikalo (1st to Last date)
    first_day_of_month = today.replace(day=1)
    last_day_of_month = first_day_of_month + relativedelta(months=1) - timedelta(days=1)

    # --- 2. Expense & Salary Calculation ---
    monthly_expenses_sum = Expense.objects.filter(
        date__gte=first_day_of_month,
        date__lte=last_day_of_month
    ).aggregate(total=Sum('amount'))['total'] or 0

    monthly_salary_paid = SalaryPayment.objects.filter(
        date__gte=first_day_of_month,
        date__lte=last_day_of_month
    ).aggregate(total=Sum('amount'))['total'] or 0

    total_outflow = monthly_expenses_sum + monthly_salary_paid

    # --- 3. Income Calculation ---
    income_entries = DailyEntry.objects.filter(
        date__gte=first_day_of_month,
        date__lte=last_day_of_month
    )
    
    total_cash = income_entries.aggregate(total=Sum('cash_received'))['total'] or 0
    total_online = income_entries.aggregate(total=Sum('online_received'))['total'] or 0
    total_income = total_cash + total_online 

    # --- 4. Net Profit ---
    net_profit = total_income - total_outflow

    # --- 5. Charts Data ---
    expense_categories = Expense.objects.filter(
        date__gte=first_day_of_month,
        date__lte=last_day_of_month
    ).values('category').annotate(sum=Sum('amount')).order_by('-sum')

    pie_labels = []
    pie_data = []
    
    for item in expense_categories:
        pie_labels.append(item['category'])
        pie_data.append(float(item['sum']))

    if monthly_salary_paid > 0:
        pie_labels.append('Staff Salary')
        pie_data.append(float(monthly_salary_paid))

    # --- Context ---
    all_canteens = Canteen.objects.all().order_by('name') 
    total_canteens = Canteen.objects.count()
    total_staff = Staff.objects.count()

    context = {
        'current_month': today.strftime("%B %Y"), # Heading ke liye (e.g. December 2025)
        'filter_date': today.strftime("%Y-%m"),   # Input box me value rakhne ke liye
        'total_canteens': total_canteens,
        'total_staff': total_staff,
        'all_canteens': all_canteens,
        'total_income': total_income,
        'total_expense': total_outflow,
        'net_profit': net_profit,
        'pie_labels': pie_labels,
        'pie_data': pie_data,
    }

    return render(request, 'management/dashboard.html', context)


# ==========================================
# 2. कैंटीन रिपोर्टिंग (Summary & Detailed)
# ==========================================
# management/views.py

# management/views.py

# management/views.py

@login_required
def canteen_summary_report(request, canteen_id):
    canteen = get_object_or_404(Canteen, pk=canteen_id)
    
    # 1. Date Filter
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    today = date.today()
    if start_date_str:
        start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
    else:
        start_date = today.replace(day=1)
        
    if end_date_str:
        end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
    else:
        end_date = today

    # 2. FETCH INCOMES & CALCULATE TOTALS (Annotate)
    # Yahan hum Python me hi total nikal rahe hain taaki HTML me {% with %} na lagana pade
    incomes = DailyEntry.objects.filter(
        canteen=canteen, 
        date__range=[start_date, end_date]
    ).annotate(
        # Income Total
        total_money=F('cash_received') + F('online_received'),
        # Food Total
        total_food=F('tea_qty') + F('nasta_qty') + F('lunch_qty') + F('dinner_qty') + 
                   F('normal_token_qty') + F('special_token_qty') + F('guest_token_qty')
    ).order_by('-date')

    # 3. FETCH EXPENSES
    raw_expenses = Expense.objects.filter(canteen=canteen, date__range=[start_date, end_date]).order_by('-date')

    # Group Expenses
    expenses_grouped = {}
    for exp in raw_expenses:
        d = exp.date
        if d not in expenses_grouped:
            expenses_grouped[d] = {'date': d, 'total_amount': 0, 'items': []}
        expenses_grouped[d]['total_amount'] += exp.amount
        expenses_grouped[d]['items'].append(exp)
    
    final_expense_list = list(expenses_grouped.values())

    # 4. OVERALL SUMMARY TOTALS
    consumption_totals = incomes.aggregate(
        total_tea=Sum('tea_qty'), total_nasta=Sum('nasta_qty'),
        total_lunch=Sum('lunch_qty'), total_dinner=Sum('dinner_qty'),
        total_normal=Sum('normal_token_qty'), total_special=Sum('special_token_qty'),
        total_guest=Sum('guest_token_qty'),
        total_cash=Sum('cash_received'), total_online=Sum('online_received')
    )

    total_income_sum = (consumption_totals['total_cash'] or 0) + (consumption_totals['total_online'] or 0)
    total_expense_sum = raw_expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    net_profit = total_income_sum - total_expense_sum

    # Token Logic
    show_tokens = False
    if canteen.billing_type and 'monthly' in str(canteen.billing_type).lower():
        show_tokens = True

    context = {
        'canteen': canteen,
        'incomes': incomes,
        'grouped_expenses': final_expense_list,
        'cons_totals': consumption_totals,
        'total_income': total_income_sum,
        'total_expense': total_expense_sum,
        'net_profit': net_profit,
        'start_date': start_date,
        'end_date': end_date,
        'show_tokens': show_tokens,
    }
    return render(request, 'management/canteen_summary_report.html', context)

@login_required
def canteen_detail_report(request, canteen_id, date_str):
    canteen = get_object_or_404(Canteen, pk=canteen_id)
    expenses = Expense.objects.filter(canteen=canteen, date=date_str).order_by('id')
    total_amount = expenses.aggregate(Sum('amount'))['amount__sum'] or 0

    context = {
        'canteen': canteen,
        'expenses': expenses,
        'total_amount': total_amount,
        'report_date': date_str,
    }
    return render(request, 'management/canteen_detail_report.html', context)


# ==========================================
# 3. पेरोल रिपोर्ट (Payroll Summary)
# ==========================================
@login_required
def payroll_summary(request):
    today = date.today()
    first_day_of_month = today.replace(day=1)
    last_day_of_month = first_day_of_month + relativedelta(months=1) - relativedelta(days=1)

    all_staff = Staff.objects.all().order_by('canteen__name', 'name')
    payroll_data = []

    for staff in all_staff:
        advance_paid = SalaryPayment.objects.filter(
            staff=staff,
            payment_type='Advance'
        ).aggregate(Sum('amount'))['amount__sum'] or 0

        paid_this_month = SalaryPayment.objects.filter(
            staff=staff,
            date__gte=first_day_of_month,
            date__lte=last_day_of_month
        ).aggregate(Sum('amount'))['amount__sum'] or 0

        leaves_this_month_periods = StaffLeave.objects.filter(
            staff=staff,
            start_date__gte=first_day_of_month,
            end_date__lte=last_day_of_month
        )
        total_leaves = sum(leave.total_days() for leave in leaves_this_month_periods)

        payroll_data.append({
            'staff': staff,
            'canteen_name': staff.canteen.name if staff.canteen else 'N/A',
            'monthly_salary': staff.monthly_salary,
            'total_advance_paid': advance_paid,
            'paid_this_month': paid_this_month,
            'leaves_this_month': total_leaves,
        })

    context = {
        'payroll_data': payroll_data,
        'current_month': today.strftime("%B %Y"),
    }
    return render(request, 'management/payroll_summary.html', context)


# ==========================================
# 4. API (JavaScript के लिए डेटा)
# ==========================================
@login_required
def get_canteen_data(request):
    canteens = Canteen.objects.all().values('id', 'billing_type')
    data = {str(c['id']): c['billing_type'] for c in canteens}
    return JsonResponse(data)


# ==========================================
# 5. स्टाफ प्रोफाइल व्यू (Bank Passbook Style)
# ==========================================
@login_required
def staff_profile_view(request, staff_id):
    staff = get_object_or_404(Staff, pk=staff_id)
    
    start_filter = request.GET.get('start_date')
    end_filter = request.GET.get('end_date')

    loop_start_date = staff.joining_date if staff.joining_date else date(date.today().year, 1, 1)
    today = date.today()

    all_transactions = []

    # 1. Payments (Debit)
    payments = SalaryPayment.objects.filter(staff=staff)
    for pay in payments:
        all_transactions.append({
            'date': pay.date,
            'description': f"Payment Taken ({pay.payment_type})",
            'credit': 0,
            'debit': pay.amount,
            'type': 'debit'
        })

    # 2. Monthly Salary (Credit) - Salary on 1st
    current_date = loop_start_date.replace(day=1)
    
    while current_date <= today.replace(day=1):
        last_day = current_date + relativedelta(months=1) - timedelta(days=1)
        
        monthly_salary = staff.monthly_salary
        leaves = StaffLeave.objects.filter(staff=staff, start_date__gte=current_date, end_date__lte=last_day)
        total_leave_days = sum(leave.total_days() for leave in leaves)
        daily_rate = monthly_salary / 30 
        deduction = daily_rate * total_leave_days
        net_salary = monthly_salary - deduction

        all_transactions.append({
            'date': current_date, 
            'description': f"Salary Credited ({current_date.strftime('%B %Y')})",
            'credit': net_salary,
            'debit': 0,
            'type': 'credit'
        })

        current_date += relativedelta(months=1)

    # 3. Sort (Oldest first for calculation)
    all_transactions.sort(key=lambda x: (x['date'], 0 if x['type'] == 'credit' else 1))

    # 4. Calculate Balance
    running_balance = 0
    final_ledger = []
    
    for trans in all_transactions:
        running_balance = running_balance + trans['credit'] - trans['debit']
        trans['balance'] = running_balance
        
        include_row = True
        if start_filter and str(trans['date']) < start_filter:
            include_row = False
        if end_filter and str(trans['date']) > end_filter:
            include_row = False
            
        if include_row:
            final_ledger.append(trans)

    # Note: final_ledger.reverse() hata diya gaya hai taaki purana upar dikhe
    
    context = {
        'staff': staff,
        'ledger_data': final_ledger,
        'overall_balance': running_balance,
        'start_date': start_filter,
        'end_date': end_filter,
    }
    return render(request, 'management/staff_profile.html', context)


# ==========================================
# 6. स्टाफ लिस्ट व्यू
# ==========================================

@login_required
def staff_list(request):
    # Sirf wahi dikhao jo abhi ACTIVE hain
    staff_members = Staff.objects.filter(is_active=True).order_by('canteen__name', 'name')
    return render(request, 'management/staff_list.html', {'all_staff': staff_members})

# ==========================================
# 7. स्टाफ डिटेल व्यू
# ==========================================
@login_required
def staff_detail_view(request, staff_id):
    staff = get_object_or_404(Staff, pk=staff_id)
    return render(request, 'management/staff_detail.html', {'staff': staff})


# ==========================================
# 8. Export Report (Download CSV)
# ==========================================
@login_required
def export_monthly_expenses(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="Monthly_Expense_Report.csv"'

    writer = csv.writer(response)
    writer.writerow(['Date', 'Category', 'Description', 'Canteen', 'Payment Mode', 'Amount'])

    today = date.today()
    first_day = today.replace(day=1)
    
    expenses = Expense.objects.filter(date__gte=first_day).order_by('-date')

    for expense in expenses:
        canteen_name = expense.canteen.name if expense.canteen else "General"
        writer.writerow([
            expense.date, 
            expense.category, 
            expense.description, 
            canteen_name, 
            expense.payment_mode, 
            expense.amount
        ])

    return response


# ==========================================
# 9. Print Staff Ledger (Print View)
# ==========================================
@login_required
def print_staff_ledger(request, staff_id):
    staff = get_object_or_404(Staff, pk=staff_id)
    loop_start_date = staff.joining_date if staff.joining_date else date(date.today().year, 1, 1)
    today = date.today()
    all_transactions = []

    payments = SalaryPayment.objects.filter(staff=staff)
    for pay in payments:
        all_transactions.append({ 'date': pay.date, 'description': f"Payment ({pay.payment_type})", 'credit': 0, 'debit': pay.amount, 'type': 'debit' })

    current_date = loop_start_date.replace(day=1)
    while current_date <= today.replace(day=1):
        last_day = current_date + relativedelta(months=1) - timedelta(days=1)
        monthly_salary = staff.monthly_salary
        leaves = StaffLeave.objects.filter(staff=staff, start_date__gte=current_date, end_date__lte=last_day)
        deduction = (monthly_salary / 30) * sum(l.total_days() for l in leaves)
        
        all_transactions.append({
            'date': current_date, # Salary on 1st
            'description': f"Salary Credited ({current_date.strftime('%B %Y')})",
            'credit': monthly_salary - deduction, 'debit': 0, 'type': 'credit'
        })
        current_date += relativedelta(months=1)

    all_transactions.sort(key=lambda x: (x['date'], 0 if x['type'] == 'credit' else 1))

    running_balance = 0
    final_ledger = []
    for trans in all_transactions:
        running_balance = running_balance + trans['credit'] - trans['debit']
        trans['balance'] = running_balance
        final_ledger.append(trans)
    
    return render(request, 'management/staff_ledger_print.html', {'staff': staff, 'ledger_data': final_ledger, 'overall_balance': running_balance, 'print_date': date.today()})


# ==========================================
# 10. Add Payment/Advance
# ==========================================
@login_required
def add_staff_payment(request, staff_id):
    staff = get_object_or_404(Staff, pk=staff_id)
    
    if request.method == 'POST':
        form = SalaryPaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.staff = staff 
            payment.save()
            messages.success(request, f"Payment of ₹{payment.amount} added for {staff.name}")
            return redirect('staff_profile', staff_id=staff.id) # Yahan galti thi, ab correct he
    else:
        form = SalaryPaymentForm(initial={'date': date.today(), 'payment_type': 'Advance'})

    return render(request, 'management/add_payment.html', {'form': form, 'staff': staff})


# ==========================================
# 11. Add Daily Expense
# ==========================================
@login_required
def add_expense(request):
    if request.method == 'POST':
        form = ExpenseForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Expense Added Successfully! ✅")
            return redirect('home_dashboard')
    else:
        form = ExpenseForm(initial={'date': date.today()})

    return render(request, 'management/add_expense.html', {'form': form})


# ==========================================
# 12. Add Daily Income (Galla)
# ==========================================
@login_required
def add_daily_entry(request):
    if request.method == 'POST':
        form = DailyEntryForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, "Daily Income Added Successfully! 💰")
                return redirect('home_dashboard')
            except:
                messages.error(request, "Error: Is Canteen aur Date ki entry pehle se hai!")
    else:
        form = DailyEntryForm(initial={'date': date.today()})

    return render(request, 'management/add_daily_entry.html', {'form': form})


# management/views.py

@login_required
def add_consumption(request):
    # 1. Canteen Types ka Data taiyar karein (JS ke liye)
    all_canteens = Canteen.objects.all()
    # Ye dictionary banayega: {1: 'Daily', 2: 'Monthly'}
    canteen_types = {c.id: c.billing_type for c in all_canteens} 

    if request.method == 'POST':
        canteen_id = request.POST.get('canteen')
        date_val = request.POST.get('date')
        
        # Check karein agar entry pehle se hai to update karein
        instance = DailyEntry.objects.filter(canteen_id=canteen_id, date=date_val).first()
        
        form = ConsumptionForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, "Data Updated Successfully! ✅")
            return redirect('home_dashboard')
    else:
        form = ConsumptionForm(initial={'date': date.today()})

    context = {
        'form': form,
        'canteen_types_json': json.dumps(canteen_types) # 👈 YE LINE BAHUT ZAROORI HAI
    }
    return render(request, 'management/add_consumption.html', context)


# management/views.py

@login_required
def consumption_report(request):
    # 1. Defaults (Aaj ka mahina)
    today = date.today()
    start_date = request.GET.get('start_date', today.replace(day=1).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', today.strftime('%Y-%m-%d'))
    selected_canteen = request.GET.get('canteen')

    # 2. Queryset (Data Filter karna)
    entries = DailyEntry.objects.filter(date__range=[start_date, end_date]).order_by('-date')

    if selected_canteen:
        entries = entries.filter(canteen_id=selected_canteen)

    # 3. Totals Calculate karna (Magic 🪄)
    totals = entries.aggregate(
        total_tea=Sum('tea_qty'),
        total_nasta=Sum('nasta_qty'),
        total_lunch=Sum('lunch_qty'),
        total_dinner=Sum('dinner_qty'),
        # Tokens
        total_normal=Sum('normal_token_qty'),
        total_special=Sum('special_token_qty'),
        total_guest=Sum('guest_token_qty')
    )

    # Context bhejna
    context = {
        'entries': entries,
        'totals': totals,
        'all_canteens': Canteen.objects.all(),
        'start_date': start_date,
        'end_date': end_date,
        'selected_canteen': int(selected_canteen) if selected_canteen else None
    }
    return render(request, 'management/consumption_report.html', context)


# management/views.py ke sabse niche paste karein

# management/views.py

@login_required
def export_canteen_excel(request, canteen_id):
    canteen = get_object_or_404(Canteen, pk=canteen_id)
    
    # 1. DATE FILTER
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    today = date.today()
    if start_date_str:
        start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
    else:
        start_date = today.replace(day=1)
        
    if end_date_str:
        end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
    else:
        end_date = today

    # 2. DATA FETCHING
    incomes = DailyEntry.objects.filter(canteen=canteen, date__range=[start_date, end_date]).order_by('date')
    expenses = Expense.objects.filter(canteen=canteen, date__range=[start_date, end_date]).order_by('date')

    # Totals Calculate
    total_income = (incomes.aggregate(Sum('cash_received'))['cash_received__sum'] or 0) + \
                   (incomes.aggregate(Sum('online_received'))['online_received__sum'] or 0)
    total_expense = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    net_profit = total_income - total_expense

    # 3. EXCEL WORKBOOK SETUP
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # --- STYLES ---
    bold_font = Font(bold=True, size=12)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") 
    center_align = Alignment(horizontal="center")
    currency_format = '#,##0'

    # --- A. REPORT HEADER ---
    ws.merge_cells('A1:G1')
    ws['A1'] = f"CANTEEN REPORT: {canteen.name.upper()}"
    ws['A1'].font = Font(bold=True, size=16, color="000000")
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:G2')
    ws['A2'] = f"Period: {start_date} to {end_date}"
    ws['A2'].alignment = center_align

    # --- B. FINANCIAL SUMMARY BOX ---
    ws['A4'] = "FINANCIAL SUMMARY"
    ws['A4'].font = bold_font
    
    headers = ["Total Income", "Total Expense", "NET PROFIT"]
    values = [total_income, total_expense, net_profit]

    ws.append([]) 
    ws.append(headers)
    ws.append(values)

    # Styling Summary
    for col_num, cell in enumerate(ws[6], 1): # Header Row
        cell.font = Font(bold=True, color="FFFFFF")
        if col_num == 1: cell.fill = PatternFill(start_color="28a745", fill_type="solid") # Green
        elif col_num == 2: cell.fill = PatternFill(start_color="dc3545", fill_type="solid") # Red
        elif col_num == 3: cell.fill = PatternFill(start_color="007bff", fill_type="solid") # Blue
        cell.alignment = center_align

    for cell in ws[7]: # Value Row
        cell.font = Font(bold=True, size=12)
        cell.number_format = currency_format
        cell.alignment = center_align

    ws.append([]) 
    ws.append([]) 

    # --- C. DAILY CONSUMPTION & INCOME ---
    ws.append(["DAILY INCOME & CONSUMPTION DETAILS"])
    ws[ws.max_row][0].font = bold_font

    columns = ['Date', 'Tea', 'Nasta', 'Lunch', 'Dinner', 'Cash (₹)', 'Online (₹)']
    
    show_tokens = False
    if canteen.billing_type and 'monthly' in str(canteen.billing_type).lower():
        show_tokens = True
        columns.extend(['Normal Token', 'Special Token', 'Guest Token'])
    
    ws.append(columns)

    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for row in incomes:
        data = [
            row.date,
            row.tea_qty, row.nasta_qty, row.lunch_qty, row.dinner_qty,
            row.cash_received, row.online_received
        ]
        if show_tokens:
            data.extend([row.normal_token_qty, row.special_token_qty, row.guest_token_qty])
        
        ws.append(data)

    ws.append([]) 
    ws.append([]) 

    # --- D. EXPENSE DETAILS ---
    ws.append(["DETAILED EXPENSES"])
    ws[ws.max_row][0].font = bold_font

    exp_columns = ['Date', 'Category', 'Description', 'Amount (₹)']
    ws.append(exp_columns)
    
    exp_header_fill = PatternFill(start_color="C0504D", fill_type="solid")
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = exp_header_fill
        cell.alignment = center_align

    for exp in expenses:
        ws.append([exp.date, exp.category, exp.description, exp.amount])

    # --- AUTO-ADJUST COLUMN WIDTH (FIXED LOGIC) ---
    # Hum enumerate use karenge taaki column index mil sake (1, 2, 3...)
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        
        # FIX: Cell se puchhne ke bajaye, hum helper function se letter nikalenge
        column_letter = get_column_letter(i) 
        
        for cell in col:
            try:
                if cell.value:
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
            except:
                pass
        
        adjusted_width = (max_length + 2)
        # Width set karte waqt sahi letter use karenge
        ws.column_dimensions[column_letter].width = adjusted_width

    # 4. RESPONSE
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"{canteen.name}_Report_{start_date}_to_{end_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
def apply_leave(request):
    if request.method == 'POST':
        form = StaffLeaveForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Leave Applied Successfully! ✅")
            return redirect('home_dashboard')
    else:
        form = StaffLeaveForm()
    
    return render(request, 'management/apply_leave.html', {'form': form})

# management/views.py

# management/views.py

@login_required
def staff_leave_history(request):
    # 1. Default: Saare leaves nikalo (Latest pehle)
    leaves = StaffLeave.objects.select_related('staff').all().order_by('-start_date')
    
    # 2. Filter Logic (Agar user ne staff select kiya hai)
    staff_id = request.GET.get('staff') # URL se ID layega
    if staff_id:
        leaves = leaves.filter(staff_id=staff_id)
    
    # 3. Dropdown ke liye saare staff members
    all_staff = Staff.objects.all()

    context = {
        'leaves': leaves,
        'all_staff': all_staff, # Dropdown list
        'selected_staff_id': int(staff_id) if staff_id else None, # Selected rakhne ke liye
        'today': date.today()
    }
    return render(request, 'management/staff_leave_history.html', context)


@login_required
def generate_payroll(request):
    # Default: Current Month
    today = date.today()
    selected_month = request.GET.get('month', today.month)
    selected_year = request.GET.get('year', today.year)
    
    selected_month = int(selected_month)
    selected_year = int(selected_year)

    # 1. Mahine ke total din nikalo (28, 30, ya 31)
    _, num_days = calendar.monthrange(selected_year, selected_month)
    
    payroll_data = []
    
    all_staff = Staff.objects.all()
    
    for staff in all_staff:
        # Salary Calculation Logic
        
        # A. Total Unpaid Leaves Count karo
        # Hum filter karenge: Staff wahi ho + Mahina/Saal wahi ho + Paid Leave FALSE ho
        unpaid_leaves_count = 0
        leaves = StaffLeave.objects.filter(
            staff=staff, 
            start_date__year=selected_year, 
            start_date__month=selected_month,
            is_paid_leave=False  # <--- SIRF UNPAID GINO
        )
        
        for leave in leaves:
            # Leave duration nikalo (End - Start + 1)
            duration = (leave.end_date - leave.start_date).days + 1
            unpaid_leaves_count += duration

        # B. Paid Leaves Count (Just for info)
        paid_leaves_count = 0
        p_leaves = StaffLeave.objects.filter(
            staff=staff, 
            start_date__year=selected_year, 
            start_date__month=selected_month,
            is_paid_leave=True   # <--- PAID VALI
        )
        for pl in p_leaves:
            duration = (pl.end_date - pl.start_date).days + 1
            paid_leaves_count += duration

        # C. Calculation 🧮
        # Per Day Salary = Monthly / Total Days (e.g. 15000 / 30 = 500)
        if num_days > 0:
            per_day_salary = staff.monthly_salary / num_days
        else:
            per_day_salary = 0
            
        deduction = unpaid_leaves_count * per_day_salary
        net_salary = staff.monthly_salary - deduction
        
        # Working Days (Present)
        present_days = num_days - unpaid_leaves_count 
        # Note: Paid leave ko hum working day hi mante hain salary ke hisab se

        payroll_data.append({
            'staff': staff,
            'total_days': num_days,
            'present_days': present_days,
            'paid_leaves': paid_leaves_count,
            'unpaid_leaves': unpaid_leaves_count,
            'per_day': round(per_day_salary, 2),
            'base_salary': staff.monthly_salary,
            'deduction': round(deduction, 2),
            'net_salary': round(net_salary, 2)
        })

    context = {
        'payroll_data': payroll_data,
        'month': selected_month,
        'year': selected_year,
        'month_name': calendar.month_name[selected_month]
    }
    return render(request, 'management/payroll_dashboard.html', context)



# management/views.py ke sabse niche


def add_staff(request):
    if request.method == 'POST':
        form = StaffForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('staff_list') # Save hone ke baad list par wapis jao
    else:
        form = StaffForm()
    
    return render(request, 'management/add_staff.html', {'form': form})

# management/views.py ke sabse niche
@login_required
def ex_staff_list(request):
    # Sirf wahi dikhao jo ACTIVE NAHI hain
    ex_staff = Staff.objects.filter(is_active=False).order_by('-leaving_date')
    return render(request, 'management/ex_staff_list.html', {'ex_staff': ex_staff})


@login_required
def mark_staff_left(request, staff_id):
    staff = get_object_or_404(Staff, id=staff_id)
    if request.method == 'POST':
        staff.is_active = False
        staff.leaving_date = timezone.now().date() # Aaj ki tarikh save karega
        staff.save()
        messages.warning(request, f"{staff.name} has been marked as Ex-Employee.")
    return redirect('staff_profile', staff_id=staff.id)